# -*- coding: utf-8 -*-
"""家庭AI小工具 — 生成测试报告 Excel"""
from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent / f"test-report-homeai-{dt.datetime.now().strftime('%Y%m%d')}-r15.xlsx"

PASS, FAIL, SKIP, WARN = "通过", "失败", "跳过", "警告"
IS_WIN = os.name == "nt"


def exists(rel: str) -> bool:
    return (ROOT / rel.replace("/", os.sep)).exists()


def read_text(rel: str) -> str:
    p = ROOT / rel.replace("/", os.sep)
    if not p.exists():
        return ""
    return p.read_text(encoding="utf-8", errors="ignore")


def contains(rel: str, pattern: str) -> bool:
    return re.search(pattern, read_text(rel), re.M) is not None


def dir_contains(rel: str, pattern: str) -> bool:
    """目录内源码是否命中正则（仅扫常见前端源文件）"""
    root = ROOT / rel.replace("/", os.sep)
    if not root.is_dir():
        return False
    rx = re.compile(pattern, re.M)
    for p in root.rglob("*"):
        if not p.is_file() or p.suffix.lower() not in {".vue", ".ts", ".js", ".tsx"}:
            continue
        try:
            if rx.search(p.read_text(encoding="utf-8", errors="ignore")):
                return True
        except Exception:
            continue
    return False


def resolve_cmd(name: str) -> str:
    """Windows 下优先用 .cmd；找不到则回退原名交给 shell。"""
    found = shutil.which(name)
    if found:
        return found
    if IS_WIN:
        found = shutil.which(f"{name}.cmd") or shutil.which(f"{name}.exe")
        if found:
            return found
    return name


def run_cmd(cmd: list[str], cwd: Path, timeout: int = 120, env: dict | None = None) -> tuple[int, str]:
    try:
        fixed = [resolve_cmd(cmd[0])] + cmd[1:]
        # Windows 上 pnpm/mvn 多为 .cmd，需 shell=True（经 COMSPEC）
        r = subprocess.run(
            fixed if not IS_WIN else subprocess.list2cmdline(fixed),
            cwd=str(cwd),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            shell=IS_WIN,
            env=env or os.environ.copy(),
        )
        out = (r.stdout or "") + "\n" + (r.stderr or "")
        return r.returncode, out
    except Exception as e:
        return 99, str(e)


def collect_rows() -> list[dict]:
    rows: list[dict] = []
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def add(
        module: str,
        case_id: str,
        name: str,
        kind: str,
        result: str,
        detail: str,
        evidence: str = "",
    ):
        rows.append(
            {
                "模块": module,
                "用例编号": case_id,
                "测试项": name,
                "类型": kind,
                "结果": result,
                "说明": detail,
                "证据/路径": evidence,
                "执行时间": now,
            }
        )

    # ---------- 自动化：管理端 Jest ----------
    vue3 = ROOT / "JeecgBoot" / "jeecgboot-vue3"
    code, out = run_cmd(
        ["pnpm", "test", "--", "tests/homeai", "tests/test.spec.ts"],
        vue3,
        timeout=180,
    )
    m = re.search(r"Tests:\s+(\d+) passed.*?(\d+) total", out, re.S)
    if m:
        passed, total = int(m.group(1)), int(m.group(2))
        jest_ok = code == 0 and passed == total and total > 0
        detail = f"Jest: {passed}/{total} passed; exit={code}"
    else:
        detail = f"Jest exit={code}; 输出摘要: {out[-300:].replace(chr(10), ' ')}"
        jest_ok = code == 0 and "PASS" in out

    add(
        "管理端-自动化",
        "FE-AUTO-001",
        "Jest 冒烟：homeai 类型 + jest 环境",
        "自动化",
        PASS if jest_ok else FAIL,
        detail,
        "JeecgBoot/jeecgboot-vue3/tests/homeai",
    )
    add(
        "管理端-自动化",
        "FE-AUTO-002",
        "HomeaiPageParams / 实体类型赋值",
        "自动化",
        PASS if jest_ok else FAIL,
        "包含在 api-types.spec.ts",
        "tests/homeai/api-types.spec.ts",
    )
    add(
        "管理端-自动化",
        "FE-AUTO-003",
        "HomeaiRecipeIngredient quantity/unit",
        "自动化",
        PASS if jest_ok else FAIL,
        "包含在 api-types.spec.ts",
        "tests/homeai/api-types.spec.ts",
    )

    # ---------- 自动化：后端 Maven compile/test（JDK17） ----------
    java_home = os.environ.get("JAVA_HOME", "")
    boot = ROOT / "JeecgBoot" / "jeecg-boot"
    if java_home and Path(java_home).exists():
        env = os.environ.copy()
        env["JAVA_HOME"] = java_home
        env["Path"] = str(Path(java_home) / "bin") + os.pathsep + env.get("Path", "")
        code, mout = run_cmd(
            [
                "mvn",
                "-B",
                "-DskipTests",
                "compile",
                "-pl",
                "jeecg-boot-module/jeecg-boot-module-homeai",
                "-am",
            ],
            boot,
            timeout=600,
            env=env,
        )
        build_ok = code == 0 or "BUILD SUCCESS" in mout
        if build_ok:
            add(
                "后端-自动化",
                "BE-AUTO-001",
                "homeai 模块 Maven 编译门禁",
                "自动化",
                PASS,
                f"BUILD SUCCESS; exit={code}；无单元测试类，compile 作为门禁",
                "jeecg-boot-module-homeai",
            )
        else:
            add(
                "后端-自动化",
                "BE-AUTO-001",
                "homeai 模块 Maven 编译门禁",
                "自动化",
                FAIL,
                f"exit={code}; " + mout[-400:].replace("\n", " "),
                "jeecg-boot-module-homeai",
            )
    else:
        add(
            "后端-自动化",
            "BE-AUTO-001",
            "homeai 模块 Maven test",
            "自动化",
            SKIP,
            "JAVA_HOME 未指向可用 JDK17，未执行",
            "",
        )

    # ---------- 静态检查：配置 / 工程化 ----------
    checks = [
        (
            "配置",
            "CFG-001",
            "管理端 docker.prod DOMAIN_URL 已配置",
            "静态",
            contains("JeecgBoot/jeecgboot-vue3/.env.docker.prod", r"VITE_GLOB_DOMAIN_URL=.+")
            and not contains("JeecgBoot/jeecgboot-vue3/.env.docker.prod", r"VITE_GLOB_DOMAIN_URL=\s*$"),
            "临时 127.0.0.1 或正式 /jeecgboot",
            ".env.docker.prod",
        ),
        (
            "配置",
            "CFG-002",
            "管理端 production DOMAIN_URL 已配置",
            "静态",
            contains("JeecgBoot/jeecgboot-vue3/.env.production", r"VITE_GLOB_DOMAIN_URL=.+")
            and not contains("JeecgBoot/jeecgboot-vue3/.env.production", r"VITE_GLOB_DOMAIN_URL=\s*$"),
            "当前临时可为 127.0.0.1；正式部署可改同域 /jeecgboot",
            ".env.production",
        ),
        (
            "配置",
            "CFG-003",
            "小程序生产 API 已配置（非 YOUR_DOMAIN 占位）",
            "静态",
            contains("JeecgUniapp/env/.env.production", r"VITE_SERVER_BASEURL\s*=")
            and not contains("JeecgUniapp/env/.env.production", r"YOUR_DOMAIN")
            and (
                contains("JeecgUniapp/env/.env.production", r"127\.0\.0\.1")
                or contains("JeecgUniapp/env/.env.production", r"https?://")
            ),
            "临时允许 127.0.0.1；上线前换真实域名",
            "JeecgUniapp/env/.env.production",
        ),
        (
            "配置",
            "CFG-003b",
            "管理端 production API_URL 为本机绝对地址",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/.env.production",
                r"VITE_GLOB_API_URL=http://127\.0\.0\.1:8080/jeecg-boot",
            ),
            "无 Nginx 时 axios 需绝对路径",
            ".env.production",
        ),
        (
            "配置",
            "CFG-004",
            "CI 含 frontend-admin 且 build:docker:prod",
            "静态",
            contains(".github/workflows/ci.yml", r"frontend-admin")
            and contains(".github/workflows/ci.yml", r"build:docker:prod"),
            "与部署文档对齐",
            ".github/workflows/ci.yml",
        ),
        (
            "配置",
            "CFG-005",
            "qiankun 默认关闭",
            "静态",
            contains("JeecgBoot/jeecgboot-vue3/.env", r"VITE_GLOB_APP_OPEN_QIANKUN\s*=\s*false"),
            "非微前端场景",
            ".env",
        ),
        (
            "配置",
            "CFG-006",
            "生产 API 临时本机约定（127.0.0.1:8080）",
            "静态",
            contains("JeecgUniapp/env/.env.production", r"127\.0\.0\.1:8080/jeecg-boot")
            and contains(
                "JeecgBoot/jeecgboot-vue3/.env.production",
                r"127\.0\.0\.1:8080/jeecg-boot",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/.env.docker.prod",
                r"127\.0\.0\.1:8080/jeecg-boot",
            ),
            "第三轮：管理端+小程序+docker.prod 均指本机",
            ".env.production / .env.docker.prod",
        ),
        (
            "管理端-功能落地",
            "FE-ST-001",
            "dynamicPages 排除 demo",
            "静态",
            contains("JeecgBoot/jeecgboot-vue3/src/utils/dynamicPages.ts", r"views/demo"),
            "减小动态页映射",
            "dynamicPages.ts",
        ),
        (
            "管理端-功能落地",
            "FE-ST-002",
            "useHomeaiCrud 存在且 planCategory 接入",
            "静态",
            exists("JeecgBoot/jeecgboot-vue3/src/views/homeai/hooks/useHomeaiCrud.ts")
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/plan/planCategory.vue",
                r"useHomeaiCrud",
            ),
            "分类页抽象",
            "hooks/useHomeaiCrud.ts",
        ),
        (
            "管理端-功能落地",
            "FE-ST-003",
            "learn/recipe/bill Category 接入 useHomeaiCrud",
            "静态",
            all(
                contains(f"JeecgBoot/jeecgboot-vue3/src/views/homeai/{p}", r"useHomeaiCrud")
                for p in (
                    "learn/learnCategory.vue",
                    "recipe/recipeCategory.vue",
                    "bill/billCategory.vue",
                )
            ),
            "第14轮",
            "views/homeai/*Category.vue",
        ),
        (
            "管理端-功能落地",
            "FE-ST-004",
            "fileList 编辑文件夹携带 visibility",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"openEditFolderModal\(\{[^}]*visibility",
            ),
            "第15轮 P0",
            "fileList.vue",
        ),
        (
            "管理端-功能落地",
            "FE-ST-005",
            "recipeApi.getById 已定义",
            "静态",
            contains("JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts", r"getById:\s*\(id"),
            "RecipeDrawer 详情",
            "api/homeai/index.ts",
        ),
        (
            "管理端-功能落地",
            "FE-ST-006",
            "SwaggerUI 动态 import",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/openapi/SwaggerUI.vue",
                r"import\('swagger-ui-dist",
            ),
            "按需加载",
            "SwaggerUI.vue",
        ),
        (
            "管理端-功能落地",
            "FE-ST-007",
            "菜单加载失败有用户提示",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/store/modules/permission.ts",
                r"菜单加载失败",
            ),
            "第13轮",
            "permission.ts",
        ),
        (
            "后端-功能落地",
            "BE-ST-001",
            "学习按分类统计 API",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/learn/controller/LearnController.java",
                r"admin/stats/category",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/service/ILearnService.java",
                r"getAdminStatsByCategory",
            ),
            "第15轮",
            "LearnController / ILearnService",
        ),
        (
            "后端-功能落地",
            "BE-ST-002",
            "homeai 模块源码目录存在",
            "静态",
            exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai"
            ),
            "业务后端根包",
            "jeecg-boot-module-homeai",
        ),
        (
            "小程序-功能落地",
            "MP-ST-001",
            "useHomeaiPageGuard 已接入主要列表页",
            "静态",
            exists("JeecgUniapp/src/pages-homeai/utils/useHomeaiPageGuard.ts")
            and all(
                contains(f"JeecgUniapp/src/{p}", r"useHomeaiPageGuard")
                for p in (
                    "pages-homeai-more/bill/index.vue",
                    "pages-homeai-more/plan/index.vue",
                    "pages-homeai-more/recipe/index.vue",
                    "pages-homeai-more/learn/index.vue",
                    "pages-homeai-more/storage/index.vue",
                )
            ),
            "第15轮",
            "useHomeaiPageGuard.ts",
        ),
        (
            "小程序-功能落地",
            "MP-ST-002",
            "菜谱食材 quantity/unit 工具存在",
            "静态",
            exists("JeecgUniapp/src/pages-homeai/utils/recipeIngredient.ts")
            and contains(
                "JeecgUniapp/src/pages-homeai/utils/recipeIngredient.ts",
                r"parseAmountToQuantityUnit",
            ),
            "第14轮",
            "recipeIngredient.ts",
        ),
        (
            "小程序-功能落地",
            "MP-ST-003",
            "学习列表跳转带 autoStart",
            "静态",
            contains(
                "JeecgUniapp/src/pages-homeai-more/learn/index.vue",
                r"autoStart=1",
            ),
            "打开即学",
            "learn/index.vue",
        ),
        (
            "小程序-功能落地",
            "MP-ST-004",
            "pages.config TabBar 主色 #1B4F8A",
            "静态",
            contains("JeecgUniapp/pages.config.ts", r"selectedColor:\s*'#1B4F8A'"),
            "第16轮视觉主色（非旧 #0960bd）",
            "pages.config.ts",
        ),
        (
            "小程序-功能落地",
            "MP-ST-008",
            "useMemberLabel 成员昵称解析",
            "静态",
            exists("JeecgUniapp/src/pages-homeai/utils/useMemberLabel.ts")
            and contains(
                "JeecgUniapp/src/pages-homeai-more/storage/StorageBrowser.vue",
                r"useMemberLabel",
            ),
            "第20轮",
            "useMemberLabel.ts",
        ),
        (
            "小程序-功能落地",
            "MP-ST-009",
            "HomeSkeleton 冷启动骨架",
            "静态",
            exists("JeecgUniapp/src/components/HomeSkeleton.vue")
            and contains(
                "JeecgUniapp/src/pages-homeai-more/storage/StorageBrowser.vue",
                r"HomeSkeleton",
            ),
            "第21轮",
            "HomeSkeleton.vue",
        ),
        (
            "小程序-功能落地",
            "MP-ST-010",
            "资料浏览器展示 thumbnailUrl",
            "静态",
            contains(
                "JeecgUniapp/src/pages-homeai-more/storage/StorageBrowser.vue",
                r"thumbnailUrl",
            ),
            "第22轮",
            "StorageBrowser.vue",
        ),
        (
            "小程序-功能落地",
            "MP-ST-011",
            "homeai-theme 含 .hai-page",
            "静态",
            contains("JeecgUniapp/src/style/homeai-theme.scss", r"\.hai-page\s*\{"),
            "第16～19轮视觉壳",
            "homeai-theme.scss",
        ),
        (
            "小程序-功能落地",
            "MP-ST-005",
            "401 跳转个人中心",
            "静态",
            contains(
                "JeecgUniapp/src/pages-homeai/api/request.ts",
                r"/pages/homeai/profile",
            ),
            "与鉴权白名单一致",
            "request.ts",
        ),
        (
            "小程序-功能落地",
            "MP-ST-006",
            "SelectUser 仅服务 Online（组件仍存在）",
            "静态",
            exists("JeecgUniapp/src/components/SelectUser/SelectUser.vue")
            and exists("JeecgUniapp/src/components/online/online-loader.vue")
            and contains(
                "JeecgUniapp/src/components/online/online-loader.vue",
                r"SelectUser",
            ),
            "R14 清理目标调整：Online 仍依赖，HomeAI 分包不引用",
            "components/SelectUser + online",
        ),
        (
            "小程序-功能落地",
            "MP-ST-007",
            "homeaiRoute 拦截 redirectTo/reLaunch",
            "静态",
            contains("JeecgUniapp/src/interceptors/homeaiRoute.ts", r"redirectTo")
            and contains("JeecgUniapp/src/interceptors/homeaiRoute.ts", r"reLaunch"),
            "深链补洞",
            "homeaiRoute.ts",
        ),
        (
            "文档",
            "DOC-001",
            "路线图含第32轮摘要",
            "静态",
            contains("docs/plan/homeai-optimization-roadmap.md", r"第 32 轮"),
            "迭代记录",
            "homeai-optimization-roadmap.md",
        ),
        (
            "后端-功能落地",
            "BE-ST-021",
            "学习提醒模板字段可配置",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/service/impl/HomeaiWxSubscribeServiceImpl.java",
                r"learn-remind-progress-field",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/service/impl/HomeaiWxSubscribeServiceImpl.java",
                r"describeLearnRemindTemplate",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/controller/HomeaiConfigController.java",
                r"wechat-learn-remind",
            ),
            "第31轮",
            "WxSubscribe / ConfigController",
        ),
        (
            "后端-功能落地",
            "BE-ST-022",
            "学习统计 Excel 导出",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/learn/controller/LearnController.java",
                r"/admin/stats/export",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/learn/learnRecord.vue",
                r"exportStats",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts",
                r"adminStatsExport",
            ),
            "第31轮",
            "learn stats export",
        ),
        (
            "后端-功能落地",
            "BE-ST-023",
            "菜谱推荐做过次数加权",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/plan/mapper/PlanInstanceMapper.java",
                r"countCompletedByRecipe",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/service/impl/RecipeServiceImpl.java",
                r"loadCookCounts",
            )
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/index.vue", r"做过多次"),
            "第32轮",
            "recommend cookCount",
        ),
        (
            "后端-功能落地",
            "BE-ST-024",
            "家庭配额运营看板 API",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/controller/HomeaiConfigController.java",
                r"/storage/families",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/service/IStorageFileService.java",
                r"listFamilyQuotaBoard",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/controller/HomeaiConfigController.java",
                r"/storage/families/batch",
            ),
            "第32轮",
            "family quota board",
        ),
        (
            "管理端-功能落地",
            "FE-ST-014",
            "家庭配额看板页 + 批量调整",
            "静态",
            exists("JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/familyQuota.vue")
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts",
                r"familyQuotaBoard",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts",
                r"batchFamilyStorageLimit",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"goFamilyQuota",
            ),
            "第32轮",
            "familyQuota.vue",
        ),
        (
            "后端-功能落地",
            "BE-ST-019",
            "家庭级存储配额覆盖",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/service/IHomeaiStorageConfigService.java",
                r"getFamilyLimitBytes",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/controller/HomeaiConfigController.java",
                r"/storage/family/",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/service/impl/StorageFileServiceImpl.java",
                r"getFamilyLimitBytes",
            ),
            "第30轮",
            "family storage override",
        ),
        (
            "后端-功能落地",
            "BE-ST-020",
            "菜谱新菜尝鲜 API",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/controller/RecipeController.java",
                r'/new"',
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/service/IRecipeService.java",
                r"listNewRecipes",
            ),
            "第30轮",
            "RecipeController /new",
        ),
        (
            "管理端-功能落地",
            "FE-ST-013",
            "家庭配额设置入口",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts",
                r"setFamilyStorageLimit",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/family/index.vue",
                r"存储配额",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"openFamilyQuota",
            ),
            "第30轮",
            "family quota UI",
        ),
        (
            "小程序-功能落地",
            "MP-ST-021",
            "菜谱首页新菜尝鲜区块",
            "静态",
            contains("JeecgUniapp/src/pages-homeai/api/recipe.ts", r"newest:")
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/index.vue", r"新菜尝鲜")
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/index.vue", r"loadNewRecipes"),
            "第30轮",
            "recipe/index.vue",
        ),
        (
            "后端-功能落地",
            "BE-ST-017",
            "学习目标 API + 提醒调度",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/learn/controller/LearnController.java",
                r'/goal"',
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/learn/service/impl/LearnRemindScheduler.java",
                r"sendLearnReminders",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/service/impl/HomeaiWxSubscribeServiceImpl.java",
                r"sendLearnRemind",
            ),
            "第29轮",
            "LearnController / Scheduler",
        ),
        (
            "后端-功能落地",
            "BE-ST-018",
            "学习多维统计 API",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/learn/controller/LearnController.java",
                r"/admin/stats/user",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/service/ILearnService.java",
                r"adminStatsByUser",
            ),
            "第29轮",
            "adminStatsByUser",
        ),
        (
            "管理端-功能落地",
            "FE-ST-012",
            "learnRecord 7/30/90 多维图表",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/learn/learnRecord.vue",
                r"近 7 日",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/learn/learnRecord.vue",
                r"adminStatsByUser",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts",
                r"adminStatsByUser",
            ),
            "第29轮",
            "learnRecord.vue",
        ),
        (
            "小程序-功能落地",
            "MP-ST-020",
            "学习每日目标 UI + 订阅授权",
            "静态",
            contains("JeecgUniapp/src/pages-homeai/api/learn.ts", r"setGoal:")
            and contains("JeecgUniapp/src/pages-homeai-more/learn/index.vue", r"今日目标")
            and contains("JeecgUniapp/src/pages-homeai-more/learn/index.vue", r"requestLearnSubscribe")
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/controller/HomeaiConfigController.java",
                r"learnRemindTemplateId",
            ),
            "第29轮",
            "learn/index.vue",
        ),
        (
            "后端-功能落地",
            "BE-ST-009",
            "菜谱 Excel 子表解析方法",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/service/IRecipeService.java",
                r"parseIngredientsFromExcel",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/controller/RecipeController.java",
                r"parseIngredientsFromExcel",
            ),
            "第24轮",
            "IRecipeService / RecipeController",
        ),
        (
            "后端-功能落地",
            "BE-ST-010",
            "用户侧回收站 /my/recycleBin",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/controller/StorageController.java",
                r"/my/recycleBin",
            ),
            "第24轮",
            "StorageController",
        ),
        (
            "小程序-功能落地",
            "MP-ST-013",
            "资料回收站页 recycle.vue",
            "静态",
            exists("JeecgUniapp/src/pages-homeai-more/storage/recycle.vue")
            and contains(
                "JeecgUniapp/src/pages-homeai-more/storage/index.vue",
                r"goRecycle",
            ),
            "第24轮",
            "storage/recycle.vue",
        ),
        (
            "配置",
            "CFG-008",
            "UniApp mp-weixin CI workflow",
            "静态",
            exists(".github/workflows/uniapp-mp-weixin.yml")
            and contains(".github/workflows/uniapp-mp-weixin.yml", r"build:mp-weixin"),
            "第24轮 nightly",
            "uniapp-mp-weixin.yml",
        ),
        (
            "文档",
            "DOC-002",
            "菜谱 Excel 导入说明含食材格式",
            "静态",
            contains("docs/guide/recipe-excel-import.md", r"名称\|数量\|单位"),
            "第24轮",
            "docs/guide/recipe-excel-import.md",
        ),
        (
            "后端-功能落地",
            "BE-ST-011",
            "AI 配额统一预检服务与 API",
            "静态",
            exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/ai/service/IHomeaiAiQuotaPrecheckService.java"
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/ai/controller/HomeaiAiQuotaController.java",
                r"/precheck",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/ai/constant/HomeaiAiQuotaScene.java",
                r"OFFICE_GENERATE",
            ),
            "第25轮",
            "HomeaiAiQuota*",
        ),
        (
            "后端-功能落地",
            "BE-ST-012",
            "Chat/Office/LLM 挂统一预检",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/ai/service/impl/HomeaiChatServiceImpl.java",
                r"assertAllowed",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/controller/StorageOfficeController.java",
                r"assertAllowed",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/ai/service/impl/HomeaiLlmServiceImpl.java",
                r"assertAllowed",
            ),
            "第25轮",
            "Chat/Office/LLM",
        ),
        (
            "小程序-功能落地",
            "MP-ST-014",
            "useHomeaiFilePick + chat/recipe/learn 接入",
            "静态",
            exists("JeecgUniapp/src/pages-homeai/utils/useHomeaiFilePick.ts")
            and contains("JeecgUniapp/src/pages-homeai-ai/ai/chat.vue", r"useHomeaiFilePick")
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/add.vue", r"useHomeaiFilePick")
            and contains("JeecgUniapp/src/pages-homeai-more/learn/add.vue", r"useHomeaiFilePick"),
            "第25轮",
            "useHomeaiFilePick.ts",
        ),
        (
            "小程序-功能落地",
            "MP-ST-015",
            "AI quotaPrecheck API",
            "静态",
            contains("JeecgUniapp/src/pages-homeai/api/ai.ts", r"quotaPrecheck")
            and contains("JeecgUniapp/src/pages-homeai-ai/ai/chat.vue", r"/ai/quota/precheck"),
            "第25轮",
            "ai.ts / chat.vue",
        ),
        (
            "后端-功能落地",
            "BE-ST-013",
            "菜谱浏览计数 + /hot",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/mapper/RecipeMapper.java",
                r"incrementViewCount",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/controller/RecipeController.java",
                r'/hot"',
            ),
            "第26轮",
            "RecipeMapper / RecipeController",
        ),
        (
            "后端-功能落地",
            "BE-ST-014",
            "综合统计 dashboard/plan-learn",
            "静态",
            exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/dashboard/controller/HomeaiDashboardController.java"
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/dashboard/controller/HomeaiDashboardController.java",
                r"plan-learn",
            ),
            "第26轮",
            "HomeaiDashboardController",
        ),
        (
            "管理端-功能落地",
            "FE-ST-012",
            "综合统计页 crossStats.vue",
            "静态",
            exists("JeecgBoot/jeecgboot-vue3/src/views/homeai/dashboard/crossStats.vue")
            and contains("JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts", r"dashboardApi")
            and exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/sql/alter_homeai_menus_iteration26.sql"
            ),
            "第26轮",
            "crossStats.vue",
        ),
        (
            "小程序-功能落地",
            "MP-ST-016",
            "菜谱热门 Tab + hot API",
            "静态",
            contains("JeecgUniapp/src/pages-homeai-more/recipe/index.vue", r"switchTab\('hot'\)")
            and contains("JeecgUniapp/src/pages-homeai/api/recipe.ts", r"hot:")
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/detail.vue", r"viewCount"),
            "第26轮",
            "recipe/index + detail",
        ),
        (
            "小程序-功能落地",
            "MP-ST-017",
            "账单导入走 useHomeaiFilePick",
            "静态",
            contains("JeecgUniapp/src/pages-homeai-more/bill/import.vue", r"useHomeaiFilePick")
            and contains("JeecgUniapp/src/pages-homeai-more/bill/import.vue", r"pickFiles"),
            "第27轮",
            "bill/import.vue",
        ),
        (
            "小程序-功能落地",
            "MP-ST-018",
            "资料上传菜单走 showStoragePickMenu",
            "静态",
            contains("JeecgUniapp/src/pages-homeai/utils/useHomeaiFilePick.ts", r"showStoragePickMenu")
            and contains(
                "JeecgUniapp/src/pages-homeai/utils/useStorageUpload.ts",
                r"showStoragePickMenu",
            ),
            "第27轮",
            "useStorageUpload.ts",
        ),
        (
            "后端-功能落地",
            "BE-ST-015",
            "家庭存储配额 + sumUsedBytesByFamily",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/dto/HomeaiStorageConfigDto.java",
                r"defaultFamilyLimitBytes",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/service/IStorageFileService.java",
                r"sumUsedBytesByFamily",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/controller/StorageController.java",
                r"perFamily",
            ),
            "第28轮",
            "storage family quota",
        ),
        (
            "后端-功能落地",
            "BE-ST-016",
            "菜谱 recommend API",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/controller/RecipeController.java",
                r"/recommend",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/recipe/service/IRecipeService.java",
                r"listRecommendRecipes",
            ),
            "第28轮",
            "RecipeController",
        ),
        (
            "小程序-功能落地",
            "MP-ST-019",
            "菜谱首页为你推荐区",
            "静态",
            contains("JeecgUniapp/src/pages-homeai/api/recipe.ts", r"recommend:")
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/index.vue", r"为你推荐")
            and contains("JeecgUniapp/src/pages-homeai-more/recipe/index.vue", r"loadRecommend"),
            "第28轮",
            "recipe/index.vue",
        ),
        (
            "管理端-功能落地",
            "FE-ST-010",
            "fileList 回收站含文件夹 Tab",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"recycleType",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"folderIds",
            ),
            "第23轮",
            "fileList.vue",
        ),
        (
            "管理端-功能落地",
            "FE-ST-011",
            "PlanDrawer 关联菜谱 recipeId",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/plan/PlanDrawer.vue",
                r"recipeId",
            ),
            "第23轮",
            "PlanDrawer.vue",
        ),
        (
            "后端-功能落地",
            "BE-ST-006",
            "文件夹回收站 deletedAt + Mapper",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/entity/StorageFolder.java",
                r"deletedAt",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/mapper/StorageFolderMapper.java",
                r"selectRecycleBinPage",
            )
            and exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/sql/alter_homeai_storage_folder_recycle.sql"
            ),
            "第23轮",
            "StorageFolder / alter SQL",
        ),
        (
            "后端-功能落地",
            "BE-ST-007",
            "存储配额配置 + 上传校验",
            "静态",
            exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/config/service/IHomeaiStorageConfigService.java"
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/service/impl/StorageFileServiceImpl.java",
                r"存储空间不足",
            ),
            "第23轮",
            "HomeaiStorageConfig / uploadFile",
        ),
        (
            "后端-功能落地",
            "BE-ST-008",
            "PlanMaster.recipeId + alter SQL",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/plan/entity/PlanMaster.java",
                r"recipeId",
            )
            and exists(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/sql/alter_homeai_plan_recipe_iteration23.sql"
            ),
            "第23轮",
            "PlanMaster / alter SQL",
        ),
        (
            "小程序-功能落地",
            "MP-ST-012",
            "首页今日下厨读取 recipeId",
            "静态",
            contains("JeecgUniapp/src/pages/homeai/index.vue", r"todayCookPlans")
            and contains("JeecgUniapp/src/pages/homeai/index.vue", r"goTodayCook"),
            "第23轮",
            "pages/homeai/index.vue",
        ),
        (
            "管理端-功能落地",
            "FE-ST-008",
            "fileList 回收站 Tab + recycleBin API",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"回收站",
            )
            and contains(
                "JeecgBoot/jeecgboot-vue3/src/api/homeai/index.ts",
                r"storage/recycleBin",
            ),
            "第22轮",
            "fileList.vue / api/homeai",
        ),
        (
            "管理端-功能落地",
            "FE-ST-009",
            "fileList 缩略图 thumbnailUrl",
            "静态",
            contains(
                "JeecgBoot/jeecgboot-vue3/src/views/homeai/storage/fileList.vue",
                r"thumbnailUrl",
            ),
            "第22轮",
            "fileList.vue",
        ),
        (
            "后端-功能落地",
            "BE-ST-003",
            "审计权限 homeai:audit:list",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/audit/controller/HomeaiAuditLogController.java",
                r"homeai:audit:list",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/sql/alter_homeai_menus_iteration22.sql",
                r"homeai:audit:list",
            ),
            "第22轮",
            "HomeaiAuditLogController / iteration22.sql",
        ),
        (
            "后端-功能落地",
            "BE-ST-004",
            "pdfbox 依赖 + StorageFile.thumbnailUrl",
            "静态",
            contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/pom.xml",
                r"<artifactId>pdfbox</artifactId>",
            )
            and contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/src/main/java/org/jeecg/modules/homeai/storage/entity/StorageFile.java",
                r"thumbnailUrl",
            ),
            "第22轮 PDF 首帧",
            "pom.xml / StorageFile.java",
        ),
        (
            "后端-功能落地",
            "BE-ST-005",
            "pom.xml 注释不含非法 --（可被 Maven 解析）",
            "静态",
            not contains(
                "JeecgBoot/jeecg-boot/jeecg-boot-module/jeecg-boot-module-homeai/pom.xml",
                r"<!--[^>]*---",
            ),
            "第三轮修复：XML 注释内禁止 --",
            "jeecg-boot-module-homeai/pom.xml",
        ),
        (
            "文档",
            "DOC-003",
            "菜谱 Excel 导入说明文档存在",
            "静态",
            exists("docs/guide/recipe-excel-import.md"),
            "第15轮起存在，第24轮已含子表格式",
            "docs/guide/recipe-excel-import.md",
        ),
        (
            "上架准备",
            "REL-001",
            "manifest targetSdkVersion >= 34",
            "静态",
            contains("JeecgUniapp/manifest.config.ts", r'targetSdkVersion["\']?\s*:\s*34')
            or contains("JeecgUniapp/src/manifest.json", r'"targetSdkVersion"\s*:\s*34'),
            "Android 上架建议",
            "manifest.config.ts",
        ),
        (
            "上架准备",
            "REL-002",
            "正式版 urlCheck 已开启（true）",
            "静态",
            contains("JeecgUniapp/manifest.config.ts", r"(?m)^\s*urlCheck:\s*true\s*,?\s*$")
            or contains("JeecgUniapp/src/manifest.json", r'(?m)^\s*"urlCheck"\s*:\s*true\s*,?\s*$'),
            "本地可保持 false；正式上架前改为 true",
            "manifest",
        ),
    ]

    for module, cid, name, kind, ok, detail, evidence in checks:
        result = PASS if ok else FAIL
        add(module, cid, name, kind, result, detail, evidence)

    # 本机后端可达性（第二轮联调）
    try:
        import urllib.request

        req = urllib.request.Request(
            "http://127.0.0.1:8080/jeecg-boot/",
            headers={"User-Agent": "homeai-test-report"},
        )
        with urllib.request.urlopen(req, timeout=3) as resp:
            status = getattr(resp, "status", 200)
            ok_probe = 200 <= int(status) < 500
            probe_detail = f"HTTP status={status}"
    except Exception as e:
        ok_probe = False
        probe_detail = f"不可达: {e}"

    add(
        "配置",
        "CFG-007",
        "本机后端 8080 可达性探测",
        "联调",
        PASS if ok_probe else WARN,
        probe_detail + "（后端未启动时为警告；临时生产 API 已指向 127.0.0.1）",
        "http://127.0.0.1:8080/jeecg-boot/",
    )

    # ---------- 运行态联调冒烟（第十三轮） ----------
    def http_json(method: str, url: str, token: str | None = None, timeout: float = 12.0):
        import json
        import urllib.error
        import urllib.request

        headers = {"User-Agent": "homeai-test-report", "Content-Type": "application/json"}
        if token:
            headers["X-Access-Token"] = token
        data = None
        if method.upper() in {"POST", "PUT"} and "?" not in url and method.upper() == "POST":
            data = b"{}"
        req = urllib.request.Request(url, data=data, headers=headers, method=method.upper())
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                status = getattr(resp, "status", 200)
                try:
                    body = json.loads(raw) if raw else {}
                except Exception:
                    body = {"_raw": raw[:200]}
                return int(status), body, None
        except urllib.error.HTTPError as e:
            raw = e.read().decode("utf-8", errors="replace") if e.fp else ""
            try:
                body = json.loads(raw) if raw else {}
            except Exception:
                body = {"_raw": raw[:200]}
            return int(e.code), body, str(e)
        except Exception as e:
            return 0, {}, str(e)

    def biz_ok(status: int, body: dict) -> bool:
        if status != 200:
            return False
        if isinstance(body, dict) and "success" in body:
            return bool(body.get("success")) and int(body.get("code") or 0) == 200
        return True

    base = "http://127.0.0.1:8080/jeecg-boot"

    # 管理端前端
    try:
        import urllib.request

        req = urllib.request.Request("http://127.0.0.1:3100/", headers={"User-Agent": "homeai-test-report"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            fe_status = getattr(resp, "status", 200)
            fe_html = resp.read().decode("utf-8", errors="ignore")
            fe_ok = 200 <= int(fe_status) < 400 and ("app" in fe_html.lower() or "vite" in fe_html.lower())
            fe_detail = f"HTTP {fe_status}"
    except Exception as e:
        fe_ok = False
        fe_detail = f"不可达: {e}"
    add(
        "联调-前端",
        "RT-FE-001",
        "管理端 Vite 3100 可达",
        "联调",
        PASS if fe_ok else WARN,
        fe_detail,
        "http://127.0.0.1:3100/",
    )

    # 小程序侧：HBuilderX 进程证据（微信开发者工具无稳定 HTTP 端口）
    try:
        import subprocess as sp

        ps = sp.run(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "Get-CimInstance Win32_Process | Where-Object { $_.Name -match 'HBuilderX|微信开发者工具|wechatdevtools' -or ($_.CommandLine -match 'HBuilderX|mp-weixin') } | Select-Object -First 1 -ExpandProperty Name",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=20,
        )
        proc_name = (ps.stdout or "").strip()
        mp_proc_ok = bool(proc_name)
        mp_proc_detail = f"检测到进程: {proc_name}" if mp_proc_ok else "未检测到 HBuilderX/微信开发者工具进程"
    except Exception as e:
        mp_proc_ok = False
        mp_proc_detail = f"进程探测失败: {e}"
    add(
        "联调-小程序",
        "RT-MP-001",
        "小程序开发工具进程在线",
        "联调",
        PASS if mp_proc_ok else WARN,
        mp_proc_detail,
        "HBuilderX / 微信开发者工具",
    )

    # 登录
    st, body, err = http_json(
        "POST",
        base + "/sys/mLogin",
    )
    # mLogin needs body
    import json
    import urllib.request

    def post_json(url: str, payload: dict, token: str | None = None):
        headers = {"User-Agent": "homeai-test-report", "Content-Type": "application/json"}
        if token:
            headers["X-Access-Token"] = token
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                return int(getattr(resp, "status", 200)), json.loads(raw)
        except Exception as e:
            return 0, {"success": False, "message": str(e)}

    st, login_body = post_json(base + "/sys/mLogin", {"username": "admin", "password": "123456"})
    admin_token = ""
    if biz_ok(st, login_body):
        admin_token = ((login_body.get("result") or {}) if isinstance(login_body.get("result"), dict) else {}).get(
            "token"
        ) or ""
        add("联调-后端", "RT-BE-001", "管理端 mLogin 获取 token", "联调", PASS, "admin 登录成功", "/sys/mLogin")
    else:
        add(
            "联调-后端",
            "RT-BE-001",
            "管理端 mLogin 获取 token",
            "联调",
            FAIL,
            f"status={st} body={str(login_body)[:180]}",
            "/sys/mLogin",
        )

    st, mp_login = post_json(base + "/homeai/user/login?code=smoke_r13_report", {})
    # login uses request param; post_json to URL with query
    import urllib.parse
    import urllib.error

    def request_raw(method: str, url: str, token: str | None = None, body: bytes | None = None):
        headers = {"User-Agent": "homeai-test-report"}
        if token:
            headers["X-Access-Token"] = token
        if body is not None:
            headers["Content-Type"] = "application/json"
        req = urllib.request.Request(url, data=body, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                return int(getattr(resp, "status", 200)), json.loads(raw) if raw else {}
        except urllib.error.HTTPError as e:
            raw = e.read().decode("utf-8", errors="replace") if e.fp else ""
            try:
                return int(e.code), json.loads(raw) if raw else {}
            except Exception:
                return int(e.code), {"success": False, "message": raw[:200]}
        except Exception as e:
            return 0, {"success": False, "message": str(e)}

    st, mp_login = request_raw("POST", base + "/homeai/user/login?code=smoke_r13_report")
    mp_token = ""
    if biz_ok(st, mp_login):
        result = mp_login.get("result") or {}
        mp_token = result.get("token") or result.get("accessToken") or ""
        add("联调-小程序", "RT-MP-002", "小程序 mock code 登录", "联调", PASS, f"openid mock 登录成功", "/homeai/user/login")
    else:
        add(
            "联调-小程序",
            "RT-MP-002",
            "小程序 mock code 登录",
            "联调",
            FAIL,
            f"status={st} msg={mp_login.get('message')}",
            "/homeai/user/login",
        )

    # API matrix
    api_cases = [
        ("联调-后端", "RT-BE-002", "学习多维统计", "/homeai/learn/admin/stats?days=7", "GET", "admin", PASS),
        ("联调-后端", "RT-BE-003", "学习趋势", "/homeai/learn/admin/stats/trend?days=7", "GET", "admin", PASS),
        ("联调-后端", "RT-BE-004", "学习提醒模板元数据", "/homeai/config/wechat-learn-remind", "GET", "admin", PASS),
        ("联调-后端", "RT-BE-005", "存储配额配置", "/homeai/config/storage", "GET", "admin", PASS),
        ("联调-后端", "RT-BE-006", "存储空间统计", "/homeai/storage/stats", "GET", "admin", PASS),
        ("联调-后端", "RT-BE-007", "综合统计 plan-learn", "/homeai/dashboard/plan-learn?days=7", "GET", "admin", WARN),
        ("联调-后端", "RT-BE-009", "家庭配额运营看板", "/homeai/config/storage/families", "GET", "admin", PASS),
        ("联调-小程序", "RT-MP-003", "学习目标查询", "/homeai/learn/goal", "GET", "mp", PASS),
        ("联调-小程序", "RT-MP-004", "学习目标设置", "/homeai/learn/goal?minutes=30", "PUT", "mp", PASS),
        ("联调-小程序", "RT-MP-005", "菜谱推荐", "/homeai/recipe/recommend?limit=3", "GET", "mp", PASS),
        ("联调-小程序", "RT-MP-006", "新菜尝鲜", "/homeai/recipe/new?limit=3", "GET", "mp", PASS),
        ("联调-小程序", "RT-MP-007", "热门菜谱", "/homeai/recipe/hot?limit=3", "GET", "mp", PASS),
        ("联调-小程序", "RT-MP-008", "微信公开配置(需登录)", "/homeai/config/wechat-public", "GET", "mp", PASS),
    ]

    for module, cid, name, path, method, who, soft in api_cases:
        token = admin_token if who == "admin" else mp_token
        if not token:
            add(module, cid, name, "联调", SKIP, "缺少 token，跳过", path)
            continue
        st, body = request_raw(method, base + path, token=token)
        ok = biz_ok(st, body)
        if ok:
            add(module, cid, name, "联调", PASS, f"success=true code={body.get('code')}", path)
        else:
            # dashboard 缺权：记警告而非失败（需执行菜单 SQL 并授权）
            if cid == "RT-BE-007" and "homeai:dashboard:view" in str(body.get("message") or ""):
                add(
                    module,
                    cid,
                    name,
                    "联调",
                    WARN,
                    f"接口可达但缺权: {body.get('message')}（请执行 alter_homeai_menus_iteration26.sql 并给 admin 勾选权限）",
                    path,
                )
            else:
                add(
                    module,
                    cid,
                    name,
                    "联调",
                    FAIL if soft == PASS else WARN,
                    f"status={st} success={body.get('success')} msg={body.get('message')}",
                    path,
                )

    # 统计导出
    if admin_token:
        try:
            req = urllib.request.Request(
                base + "/homeai/learn/admin/stats/export?days=7",
                headers={"User-Agent": "homeai-test-report", "X-Access-Token": admin_token},
                method="GET",
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = resp.read()
                ctype = resp.headers.get("Content-Type", "")
                ok_export = len(data) > 1000 and ("sheet" in ctype or "octet" in ctype or "excel" in ctype)
                add(
                    "联调-后端",
                    "RT-BE-008",
                    "学习统计 Excel 导出",
                    "联调",
                    PASS if ok_export else FAIL,
                    f"bytes={len(data)} content-type={ctype}",
                    "/homeai/learn/admin/stats/export",
                )
        except Exception as e:
            add("联调-后端", "RT-BE-008", "学习统计 Excel 导出", "联调", FAIL, str(e), "/homeai/learn/admin/stats/export")
    else:
        add("联调-后端", "RT-BE-008", "学习统计 Excel 导出", "联调", SKIP, "无 admin token", "/homeai/learn/admin/stats/export")

    # 匿名 wechat-public：应公开可访问
    st, body = request_raw("GET", base + "/homeai/config/wechat-public")
    if biz_ok(st, body):
        add("联调-后端", "RT-BE-009", "wechat-public 匿名访问", "联调", PASS, "匿名可访问", "/homeai/config/wechat-public")
    elif st == 401:
        add(
            "联调-后端",
            "RT-BE-009",
            "wechat-public 匿名访问",
            "联调",
            FAIL,
            "未登录仍返回 401（应加入 PUBLIC_PATHS）",
            "/homeai/config/wechat-public",
        )
    else:
        add(
            "联调-后端",
            "RT-BE-009",
            "wechat-public 匿名访问",
            "联调",
            WARN,
            f"status={st} msg={body.get('message')}",
            "/homeai/config/wechat-public",
        )

    return rows


def style_workbook(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "测试结果明细"

    headers = ["模块", "用例编号", "测试项", "类型", "结果", "说明", "证据/路径", "执行时间"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0960BD")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    fills = {
        PASS: PatternFill("solid", fgColor="C6EFCE"),
        FAIL: PatternFill("solid", fgColor="FFC7CE"),
        SKIP: PatternFill("solid", fgColor="FFEB9C"),
        WARN: PatternFill("solid", fgColor="FCE4D6"),
    }

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for r in rows:
        ws.append([r[h] for h in headers])
        row_idx = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if headers[col - 1] == "结果":
                cell.fill = fills.get(r["结果"], PatternFill())
                cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [16, 14, 42, 10, 8, 48, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    ws.freeze_panes = "A2"

    # 汇总 sheet
    summary = wb.create_sheet("汇总", 0)
    total = len(rows)
    counts = {PASS: 0, FAIL: 0, SKIP: 0, WARN: 0}
    for r in rows:
        counts[r["结果"]] = counts.get(r["结果"], 0) + 1

    summary.append(["家庭AI小工具 — 测试报告汇总"])
    summary["A1"].font = Font(bold=True, size=14, color="0960BD")
    summary.append(["生成时间", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["项目路径", str(ROOT)])
    summary.append([])
    summary.append(["结果", "数量", "占比"])
    for k in (PASS, FAIL, WARN, SKIP):
        pct = f"{(counts[k] / total * 100):.1f}%" if total else "0%"
        summary.append([k, counts[k], pct])
        summary.cell(summary.max_row, 1).fill = fills[k]
    summary.append([])
    summary.append(["合计", total, "100%"])
    summary.append([])
    summary.append(["说明"])
    summary.append(["通过", "自动化执行成功或静态检查命中预期"])
    summary.append(["失败", "自动化失败或关键文件/配置缺失"])
    summary.append(["警告", "未阻塞开发但上线前必须处理（如生产域名占位、urlCheck）"])
    summary.append(["跳过", "环境不满足未能执行（如无测试类/JDK）"])
    summary.append([])
    summary.append(["未覆盖说明"])
    summary.append(
        [
            "本报告不含 E2E/真机联调（需后端服务、微信开发者工具、真实域名）。"
            "建议在填实 YOUR_DOMAIN 后补充接口冒烟与小程序真机用例。"
        ]
    )

    for col in range(1, 4):
        summary.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 60


def main() -> int:
    # 确保 JAVA_HOME
    if not os.environ.get("JAVA_HOME"):
        for cand in (
            r"C:\Program Files\Java\jdk-17",
            r"C:\Users\57089\.jdks\ms-17.0.19",
        ):
            if Path(cand).exists():
                os.environ["JAVA_HOME"] = cand
                os.environ["Path"] = str(Path(cand) / "bin") + os.pathsep + os.environ.get("Path", "")
                break

    rows = collect_rows()
    wb = Workbook()
    # openpyxl 默认有一 sheet，先写明细再插入汇总
    style_workbook(wb, rows)
    # style_workbook 把汇总插到 index0，明细改名为测试结果明细
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"REPORT={OUT}")
    counts = {}
    for r in rows:
        counts[r["结果"]] = counts.get(r["结果"], 0) + 1
    print("COUNTS=" + ",".join(f"{k}:{v}" for k, v in counts.items()))
    return 0 if counts.get(FAIL, 0) == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
