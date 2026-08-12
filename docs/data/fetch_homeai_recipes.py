# -*- coding: utf-8 -*-
"""
从开源项目 Anduin2017/HowToCook 解析常用菜谱，导出为 HomeAI Excel 导入格式。

数据源：https://github.com/Anduin2017/HowToCook （社区开源菜谱）
输出列对齐 docs/guide/recipe-excel-import.md
"""

from __future__ import annotations

import argparse
import io
import re
import ssl
import zipfile
from pathlib import Path
from typing import Optional

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SRC = ROOT / "tmp" / "HowToCook-master" / "dishes"
DEFAULT_OUT = Path(__file__).resolve().parent / "homeai-recipes-import.xlsx"
ZIP_URL = "https://codeload.github.com/Anduin2017/HowToCook/zip/refs/heads/master"

# HowToCook 目录 -> 系统分类 ID（init_homeai_recipe_category.sql）
CATEGORY_MAP = {
    "vegetable_dish": "rc_hot",
    "meat_dish": "rc_hot",
    "aquatic": "rc_hot",
    "soup": "rc_soup",
    "staple": "rc_staple",
    "breakfast": "rc_staple",
    "dessert": "rc_bake",
    "drink": "rc_drink",
    "condiment": "rc_other",
    "semi-finished": "rc_other",
}

# 优先收录的家常菜（按常见程度）
PRIORITY_NAMES = [
    "宫保鸡丁", "鱼香肉丝", "麻婆豆腐", "回锅肉", "红烧肉", "糖醋排骨", "可乐鸡翅",
    "西红柿炒蛋", "番茄炒蛋", "酸辣土豆丝", "蒜蓉西兰花", "清蒸鲈鱼", "水煮肉片",
    "水煮鱼", "干煸豆角", "地三鲜", "鱼香茄子", "红烧茄子", "青椒肉丝", "土豆炖牛肉",
    "红烧排骨", "糖醋里脊", "京酱肉丝", "蒜苔炒肉", "芹菜炒肉", "洋葱炒肉", "木须肉",
    "蚂蚁上树", "酸菜鱼", "番茄鸡蛋汤", "紫菜蛋花汤", "冬瓜排骨汤", "玉米排骨汤",
    "蛋炒饭", "扬州炒饭", "炒面", "葱油拌面", "炸酱面", "馄饨", "饺子", "包子",
    "豆浆", "油条", "煎饼", "茶叶蛋", "卤蛋", "皮蛋瘦肉粥", "南瓜粥", "小米粥",
    "红烧鸡翅", "辣子鸡", "口水鸡", "白切鸡", "啤酒鸭", "可乐鸡", "咖喱鸡",
    "红烧豆腐", "家常豆腐", "葱爆羊肉", "孜然羊肉", "西红柿牛腩", "胡萝卜炖牛腩",
    "蒜蓉虾", "油焖大虾", "清炒虾仁", "干锅花菜", "蚝油生菜", "炒青菜", "上汤娃娃菜",
    "醋溜白菜", "凉拌黄瓜", "凉拌木耳", "拍黄瓜", "凉拌海带丝", "凉拌豆腐",
    "蛋黄酥", "戚风蛋糕", "提拉米苏", "烤红薯", "烤鸡翅", "土豆泥", "沙拉",
    "奶茶", "柠檬水", "酸梅汤", "绿豆汤", "银耳莲子汤", "红豆沙",
    "麻辣香锅", "干锅土豆片", "香菇滑鸡", "啤酒鸡翅", "蜜汁鸡翅", "红烧狮子头",
    "锅包肉", "东北乱炖", "酱牛肉", "卤牛肉", "炖排骨", "清蒸鸡蛋", "蒸蛋羹",
    "番茄炒蛋", "韭菜炒蛋", "青椒炒蛋", "火腿炒蛋", "虾仁炒蛋", "木耳炒鸡蛋",
    "蒜蓉粉丝蒸虾", "清蒸螃蟹", "香辣小龙虾", "蒜蓉扇贝", "豆腐脑", "胡辣汤",
    "牛肉面", "兰州拉面", "担担面", "热干面", "刀削面", "凉皮", "肉夹馍",
    "煎饼果子", "手抓饼", "葱油饼", "烙饼", "馒头", "花卷", "烧麦", "小笼包",
]

COLD_KEYWORDS = ("凉拌", "拍黄", "白切", "口水", "凉皮", "沙拉")
HEADER = [
    "菜谱名称",
    "分类ID",
    "难度(1-5)",
    "用时(分钟)",
    "份数",
    "食材(名称|数量|单位;...)",
    "步骤(分号分隔)",
    "小贴士",
    "可见性(private/family/public)",
]


def download_source(dest_root: Path) -> Path:
    dest_root.mkdir(parents=True, exist_ok=True)
    dishes = dest_root / "HowToCook-master" / "dishes"
    if dishes.exists() and any(dishes.rglob("*.md")):
        return dishes
    print("下载 HowToCook 仓库 zip ...")
    ctx = ssl.create_default_context()
    try:
        import certifi

        ctx = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        pass
    import urllib.request

    req = urllib.request.Request(ZIP_URL, headers={"User-Agent": "Mozilla/5.0 HomeAI-RecipeFetcher"})
    with urllib.request.urlopen(req, context=ctx, timeout=180) as resp:
        data = resp.read()
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        zf.extractall(dest_root)
    if not dishes.exists():
        raise RuntimeError(f"解压后未找到 dishes 目录: {dishes}")
    return dishes


def category_id(folder: str, name: str) -> str:
    if any(k in name for k in COLD_KEYWORDS):
        return "rc_cold"
    return CATEGORY_MAP.get(folder, "rc_other")


def parse_difficulty(text: str) -> int:
    m = re.search(r"预估烹饪难度：([★☆]+)", text)
    if not m:
        return 3
    stars = m.group(1).count("★")
    return max(1, min(5, stars or 3))


def parse_cook_time(text: str, difficulty: int) -> int:
    patterns = [
        r"(?:大约|约|需要|耗时)?\s*(\d+(?:\.\d+)?)\s*小时",
        r"(?:大约|约|需要|耗时)?\s*(\d+)\s*分钟",
        r"(\d+)\s*-\s*(\d+)\s*分钟",
        r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)\s*小时",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if not m:
            continue
        if "小时" in pat:
            if m.lastindex == 2:
                hours = (float(m.group(1)) + float(m.group(2))) / 2
            else:
                hours = float(m.group(1))
            return max(10, int(round(hours * 60)))
        if m.lastindex == 2:
            return max(5, (int(m.group(1)) + int(m.group(2))) // 2)
        return max(5, int(m.group(1)))
    # 按难度估算
    return {1: 15, 2: 25, 3: 40, 4: 60, 5: 90}.get(difficulty, 40)


def parse_servings(text: str) -> int:
    patterns = [
        r"一份正好够\s*(\d+)\s*个人",
        r"默认一人",
        r"两人?也够吃",
        r"可供\s*(\d+)\s*人",
        r"(\d+)\s*人份",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if not m:
            continue
        if "默认一人" in pat:
            return 1
        if "两人" in pat:
            return 2
        return max(1, int(m.group(1)))
    return 2


CN_NUM = {
    "一": "1",
    "二": "2",
    "两": "2",
    "三": "3",
    "四": "4",
    "五": "5",
    "六": "6",
    "七": "7",
    "八": "8",
    "九": "9",
    "十": "10",
}

# 常见量词（用于「1 盒豆腐」「2 片生姜」）
UNIT_WORDS = (
    "盒", "枚", "个", "只", "条", "根", "片", "瓣", "颗", "克", "千克", "斤", "两",
    "毫升", "升", "勺", "汤匙", "茶匙", "杯", "碗", "把", "块", "段", "滴", "适量",
    "g", "kg", "ml", "L", "l",
)

SKIP_ING_PREFIX = (
    "注意",
    "每份",
    "每次",
    "按照",
    "使用上述",
    "计算出",
    "必须配料",
    "进阶配料",
    "可选配料",
    "可选原料",
    "一份",
)

# 饮品优先家常中式；排除鸡尾酒等
DRINK_PRIORITY = ("豆浆", "酸梅汤", "柠檬水", "奶茶", "绿豆汤", "可乐", "咖啡", "茶", "杨枝甘露", "酸梅")
DRINK_EXCLUDE = ("B52", "Mojito", "鸡尾酒", "长岛", "龙舌兰", "威士忌", "马天尼")


def _normalize_cn_qty(text: str) -> str:
    """仅当「两瓣」「一盒」这类数量+量词时转换，避免误伤「五花肉」."""
    for k, v in CN_NUM.items():
        for u in sorted(UNIT_WORDS, key=len, reverse=True):
            prefix = k + u
            if text.startswith(prefix):
                return v + text[len(k) :]
            prefix2 = k + " " + u
            if text.startswith(prefix2):
                return v + text[len(k) :]
    return text


def _split_qty_unit(raw: str) -> tuple[str, str]:
    raw = raw.strip()
    if not raw:
        return "", ""
    raw = _normalize_cn_qty(raw)
    # 约 350g / 115g / 2 个 / 10-15ml / 1 支（约 350g）
    m = re.match(
        r"^(?:约\s*)?(\d+(?:\.\d+)?(?:\s*[-~～]\s*\d+(?:\.\d+)?)?)\s*([a-zA-Z\u4e00-\u9fff%]+)?",
        raw,
    )
    if not m:
        return "", raw
    qty = m.group(1).replace(" ", "").replace("～", "-").replace("~", "-")
    if "-" in qty:
        qty = qty.split("-", 1)[0]
    unit = (m.group(2) or "").strip()
    unit = re.split(r"[（(]", unit, maxsplit=1)[0].strip()
    if len(unit) > 12:
        unit = unit[:12]
    return qty, unit


def parse_ingredient_line(line: str, allow_name_only: bool = False) -> Optional[str]:
    line = line.strip()
    if not line or line.startswith("#"):
        return None
    line = re.sub(r"^[-*+]\s*", "", line)
    line = re.sub(r"\s*约\s*(?=\d)", " ", line)
    line = re.sub(r"\s+", " ", line).strip()
    if any(line.startswith(p) for p in SKIP_ING_PREFIX):
        return None
    if "分量" in line and ("杯" in line or "毫升" in line or "约" in line):
        return None

    # 「五花肉的用量为 0.5 斤/...」
    m = re.match(r"^(.+?)的用量为\s*(.+)$", line)
    if m:
        name = re.split(r"[（(]", m.group(1).strip(), maxsplit=1)[0].strip()
        qty, unit = _split_qty_unit(m.group(2))
        if name:
            return f"{name}|{qty}|{unit}"

    # 名称 = 数量单位
    if "=" in line:
        name, rest = line.split("=", 1)
        name = re.sub(r"^[-*+\d\\.、)\s]+", "", name).strip()
        name = re.sub(r"的用量为?$", "", name).strip()
        name = re.split(r"[（(]", name, maxsplit=1)[0].strip()
        qty, unit = _split_qty_unit(rest)
        if not name or len(name) > 40 or name in ("用量", "份量"):
            return None
        return f"{name}|{qty}|{unit}"

    # 「20-30g 五花肉」「5g 酱油」「1 盒内脂豆腐」「2 片生姜」「两瓣大蒜」
    s = _normalize_cn_qty(line)
    m = re.match(
        r"^(?:约\s*)?(\d+(?:\.\d+)?(?:\s*[-~～]\s*\d+(?:\.\d+)?)?)\s*"
        r"([a-zA-Z]+|盒|枚|个|只|条|根|片|瓣|颗|克|千克|斤|两|毫升|升|勺|汤匙|茶匙|杯|碗|把|块|段|滴)?\s*"
        r"(.+)$",
        s,
    )
    if m:
        qty = m.group(1).replace(" ", "").replace("～", "-").replace("~", "-")
        if "-" in qty:
            qty = qty.split("-", 1)[0]
        unit = (m.group(2) or "").strip()
        name = m.group(3).strip()
        name = re.split(r"[（(]", name, maxsplit=1)[0].strip()
        # 若单位空且名称以常见单位开头，拆出来
        if not unit:
            for u in sorted(UNIT_WORDS, key=len, reverse=True):
                if name.startswith(u) and len(name) > len(u):
                    unit, name = u, name[len(u) :].strip()
                    break
        if name and 1 <= len(name) <= 40 and not name.startswith("的"):
            # 过滤明显不是食材名的解析结果
            if "用量" in name or "份量" in name or "分量" in name or "/" in name:
                # 尝试从「五花肉的用量为 0.5 斤」提取名称
                nm = re.split(r"的用量", name, maxsplit=1)[0].strip()
                if nm and "用量" not in nm and "份量" not in nm:
                    return f"{nm}|{qty}|{unit}"
                return None
            return f"{name}|{qty}|{unit}"

    # 名称 在前：土豆 2 个
    m = re.match(
        r"^([^\d=]{1,40}?)\s+(\d+(?:\.\d+)?(?:\s*[-~～]\s*\d+(?:\.\d+)?)?\s*[a-zA-Z\u4e00-\u9fff%].*)$",
        line,
    )
    if m:
        name = re.split(r"[（(]", m.group(1).strip(), maxsplit=1)[0].strip()
        name = re.sub(r"的用量为?$", "", name).strip()
        qty, unit = _split_qty_unit(m.group(2))
        if name and len(unit) <= 12 and name not in ("用量", "份量"):
            return f"{name}|{qty}|{unit}"

    if not allow_name_only:
        return None
    name = re.split(r"[（(]", line, maxsplit=1)[0].strip()
    name = re.sub(r"^[-*+\d\\.、)\s]+", "", name)
    if name in ("水果刀", "菜刀", "碗", "锅", "铲子", "保鲜膜"):
        return None
    if (
        1 <= len(name) <= 20
        and not re.search(r"[。；;=]", name)
        and not re.match(r"^\d", name)
    ):
        return f"{name}||"
    return None


def extract_section(text: str, title: str) -> str:
    pat = rf"^##\s*{re.escape(title)}\s*$"
    m = re.search(pat, text, flags=re.M)
    if not m:
        return ""
    start = m.end()
    nxt = re.search(r"^##\s+", text[start:], flags=re.M)
    end = start + nxt.start() if nxt else len(text)
    return text[start:end].strip()


def parse_ingredients(text: str) -> str:
    calc = extract_section(text, "计算")
    tools = extract_section(text, "必备原料和工具")
    items: list[str] = []
    seen = set()

    def _absorb(block: str, allow_name_only: bool) -> None:
        for raw in block.splitlines():
            if re.match(r"^#{1,4}\s", raw):
                continue
            parsed = parse_ingredient_line(raw, allow_name_only=allow_name_only)
            if not parsed:
                continue
            # 补充原料清单名称时做简单去重（忽略空格/约）
            key = re.sub(r"\s+|约", "", parsed.split("|", 1)[0])
            if key in seen:
                continue
            seen.add(key)
            # 也登记原始名
            seen.add(parsed.split("|", 1)[0])
            items.append(parsed)
            if len(items) >= 18:
                return

    # 优先解析「计算」区（带数量）；再从原料清单补名称（如主料鲈鱼）
    _absorb(calc, allow_name_only=False)
    if len(items) < 3:
        _absorb(calc, allow_name_only=True)
    _absorb(tools, allow_name_only=True)
    return ";".join(items)


def parse_steps(text: str) -> str:
    ops = extract_section(text, "操作")
    if not ops:
        return ""
    # 若有多版本，优先取第一个含编号步骤的子块
    blocks = re.split(r"^###\s+.+$", ops, flags=re.M)
    candidate = ops
    for b in blocks:
        if re.search(r"^\s*\d+[\\.、)]\s*", b, flags=re.M):
            candidate = b
            break
    steps: list[str] = []
    for raw in candidate.splitlines():
        line = raw.strip()
        m = re.match(r"^(\d+)[\\.、)]\s*(.+)$", line)
        if not m:
            continue
        desc = m.group(2).strip()
        desc = re.sub(r"\*+", "", desc)
        desc = re.sub(r"\s+", " ", desc)
        if desc:
            steps.append(desc)
    return ";".join(steps[:15])


def parse_tips(text: str) -> str:
    extra = extract_section(text, "附加内容")
    tips: list[str] = []
    for raw in extra.splitlines():
        line = re.sub(r"^[-*+]\s*", "", raw.strip())
        if not line or line.startswith("如果您遵循") or line.startswith("参考资料") or line.startswith("["):
            continue
        if line.startswith("http") or line.startswith("!"):
            continue
        tips.append(line)
        if len(tips) >= 3:
            break
    return "；".join(tips)[:500]


def recipe_name_from_file(path: Path, text: str) -> str:
    m = re.search(r"^#\s*(.+?)的做法\s*$", text, flags=re.M)
    if m:
        return m.group(1).strip()
    return path.stem


def parse_recipe(path: Path, folder: str) -> Optional[dict]:
    text = path.read_text(encoding="utf-8")
    name = recipe_name_from_file(path, text)
    if name in ("示例菜", "模板"):
        return None
    difficulty = parse_difficulty(text)
    ingredients = parse_ingredients(text)
    steps = parse_steps(text)
    if not ingredients and not steps:
        return None
    return {
        "name": name,
        "categoryId": category_id(folder, name),
        "difficulty": difficulty,
        "cookTime": parse_cook_time(text, difficulty),
        "servings": parse_servings(text),
        "ingredients": ingredients,
        "steps": steps,
        "tips": parse_tips(text),
        "visibility": "public",
        "folder": folder,
        "source": str(path),
    }


def collect_recipes(dishes_dir: Path, limit: int = 100) -> list[dict]:
    all_md = [p for p in dishes_dir.rglob("*.md") if "template" not in p.parts]
    by_name: dict[str, Path] = {}
    for p in all_md:
        key = p.stem
        prev = by_name.get(key)
        if prev is None or p.parent.name == p.stem:
            by_name[key] = p

    selected: list[dict] = []
    used = set()

    def try_add(name: str, path: Path) -> bool:
        if name in used or len(selected) >= limit:
            return False
        folder = path.relative_to(dishes_dir).parts[0]
        if folder in ("condiment", "semi-finished", "template"):
            return False
        if folder == "drink" and any(x in name for x in DRINK_EXCLUDE):
            return False
        if folder == "drink" and re.search(r"[A-Za-z]", name) and not any(k in name for k in DRINK_PRIORITY):
            return False
        recipe = parse_recipe(path, folder)
        if not recipe:
            return False
        # 至少要有步骤，且食材不少于 2 项（家常菜可用性）
        if not recipe["steps"] or recipe["ingredients"].count(";") < 1:
            return False
        selected.append(recipe)
        used.add(name)
        return True

    # 饮品优先中式家常（放在家常菜之后补齐）
    drink_paths = []
    for name in DRINK_PRIORITY:
        path = by_name.get(name)
        if path:
            drink_paths.append((name, path))

    for name in PRIORITY_NAMES:
        if len(selected) >= limit:
            break
        path = by_name.get(name)
        if path:
            try_add(name, path)

    for name, path in drink_paths:
        if len(selected) >= limit:
            break
        try_add(name, path)

    quotas = {
        "meat_dish": 25,
        "vegetable_dish": 25,
        "aquatic": 10,
        "soup": 10,
        "staple": 12,
        "breakfast": 8,
        "dessert": 5,
        "drink": 5,
    }
    folder_counts: dict[str, int] = {}
    for r in selected:
        folder_counts[r["folder"]] = folder_counts.get(r["folder"], 0) + 1

    remaining = []
    for name, path in sorted(by_name.items(), key=lambda x: x[0]):
        if name in used:
            continue
        folder = path.relative_to(dishes_dir).parts[0]
        if folder not in quotas:
            continue
        remaining.append((folder, name, path))

    for folder, max_n in quotas.items():
        if len(selected) >= limit:
            break
        need = max_n - folder_counts.get(folder, 0)
        if need <= 0:
            continue
        for f, name, path in remaining:
            if f != folder:
                continue
            if try_add(name, path):
                folder_counts[folder] = folder_counts.get(folder, 0) + 1
                need -= 1
                if need <= 0 or len(selected) >= limit:
                    break

    if len(selected) < limit:
        for f, name, path in remaining:
            if len(selected) >= limit:
                break
            try_add(name, path)

    return selected[:limit]


def write_excel(rows: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "导出信息"
    # 对齐 AutoPoi: titleRows=2, headRows=1
    ws.append(["菜谱列表数据"])
    ws.append(["导出人:系统", "数据来源:Anduin2017/HowToCook"])
    ws.append(HEADER)
    for r in rows:
        ws.append(
            [
                r["name"],
                r["categoryId"],
                r["difficulty"],
                r["cookTime"],
                r["servings"],
                r["ingredients"],
                r["steps"],
                r["tips"],
                r["visibility"],
            ]
        )
    from openpyxl.utils import get_column_letter

    widths = [16, 12, 10, 12, 8, 50, 60, 30, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="拉取 HowToCook 并导出 HomeAI 菜谱 Excel")
    parser.add_argument("--limit", type=int, default=100, help="导出条数，默认 100")
    parser.add_argument("--src", type=Path, default=DEFAULT_SRC, help="本地 dishes 目录")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT, help="输出 xlsx 路径")
    parser.add_argument("--download", action="store_true", help="强制重新下载源数据")
    args = parser.parse_args()

    dishes = args.src
    if args.download or not dishes.exists():
        dishes = download_source(ROOT / "tmp")

    rows = collect_recipes(dishes, limit=args.limit)
    write_excel(rows, args.out)

    # 统计
    from collections import Counter

    cat = Counter(r["categoryId"] for r in rows)
    folder = Counter(r["folder"] for r in rows)
    with_ing = sum(1 for r in rows if r["ingredients"])
    with_steps = sum(1 for r in rows if r["steps"])
    print(f"已导出 {len(rows)} 条 -> {args.out}")
    print(f"分类分布: {dict(cat)}")
    print(f"来源目录: {dict(folder)}")
    print(f"含食材 {with_ing} / 含步骤 {with_steps}")
    print("示例:")
    for r in rows[:5]:
        print(f"  - {r['name']} [{r['categoryId']}] 难度{r['difficulty']} {r['cookTime']}分钟")


if __name__ == "__main__":
    main()
